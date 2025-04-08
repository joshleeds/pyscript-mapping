import pandas as pd
import numpy as np
from js import document, FileReader, console, Blob, URL, Uint8Array,window 
from io import StringIO, BytesIO
from pyodide.ffi import create_proxy
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import warnings

if window.location.hostname in ["localhost", "127.0.0.1"]:
    # Development mode
    warnings.filterwarnings("default")
else:
    # Production mode (e.g. on vercel or other domains)
    warnings.filterwarnings("ignore")

console.log("Script loaded and running...")

def allow_click(event):
    file_input = document.getElementById("excel-upload")
    process_btn = document.getElementById("process-btn")

    if file_input.files.length > 0:
        process_btn.disabled = False
        process_btn.classList.remove("bg-gray-300", "text-gray-500", "cursor-not-allowed")
        process_btn.classList.add("bg-blue-600", "hover:bg-blue-700", "text-white")
    else:
        process_btn.disabled = True
        process_btn.classList.remove("bg-blue-600", "hover:bg-blue-700", "text-white")
        process_btn.classList.add("bg-gray-300", "text-gray-500", "cursor-not-allowed")

def remove_file(event):
    file_input = document.getElementById("excel-upload")
    remove_button = document.getElementById("remove-file-btn")
    file_input.value = ""
    process_btn = document.getElementById("process-btn")
    process_btn.disabled = True
    process_btn.classList.remove("bg-blue-600", "hover:bg-blue-700", "text-white")
    process_btn.classList.add("bg-gray-300", "text-gray-500", "cursor-not-allowed")


def handle_upload(event):
    console.log("Uploading file...")
    file = document.getElementById('excel-upload').files.item(0)
    reader = FileReader.new()

    def combine(event):
        file_content = event.target.result
        
        # reads the file as either a CSV or Excel file depending on the file extension most likely will be a Excel
        if file.name.endswith('.csv'):
             df = pd.read_csv(StringIO(file_content.decode('utf-8')))
        elif file.name.endswith(('.xls', '.xlsx')):
            binary_data = memoryview(file_content.to_py())  # convert JsProxy to Python bytes-like object
            excel_io = BytesIO(binary_data)
            sheet_names = pd.ExcelFile(excel_io).sheet_names
            ##Pick the sheet name 
            if "Transactions" in sheet_names:
                read_sheet = "Transactions"
            else:
                read_sheet = sheet_names[0]
            console.log(f"Reading sheet: {read_sheet}")
            excel_io.seek(0)  # Reset the pointer to the start of the BytesIO object 
            df = pd.read_excel(excel_io, sheet_name=0)
        else:
            raise ValueError("Unsupported file format. Please upload a CSV or Excel file.")
        
        ###Combining of the data
        console.log("Data read")
        # Keep only the specified columns
        # required_columns = ['Date', 'Description', 'Paid Out', 'Balance', 'Classification']
        # df = df[required_columns]
        training_data = df[df['Classification'].notnull()]
        training_data = training_data.dropna()

        class_map = dict(zip(training_data['Description'], training_data['Classification']))
        df['Classification'] = df['Classification'].fillna(df['Description'].map(class_map))

        count_null = df['Classification'].isnull().sum()
        console.log(f"Number of null values in Classification column: {count_null}")
        console.log(f"Number of rows in the data: {df.shape[0]}")

        ##Highlight the null values in the Classification column
        def highlight_and_style(df):
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            wb = Workbook()
            ws = wb.active
            ws.title = "Transactions"
            ws.append(list(df.columns))

            for row in df.itertuples(index=False):
                ws.append(row)
                if pd.isnull(row.Classification):
                    for col in range(1, len(df.columns) + 1):
                        ws.cell(row=ws.max_row, column=col).fill = yellow_fill

            # Adjust column widths
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length

            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
        
        output_data = highlight_and_style(df)

        # Convert the DataFrame to a excel string
        # excel_buffer = BytesIO()
        # df.to_excel(excel_buffer, index=False, engine='openpyxl')
        # output_data = excel_buffer.getvalue()  # Get the binary content
        console.log(f"Excel data size: {len(output_data)} bytes")
        
        # Create a Blob and a downloadable link
        excel_uint8 = Uint8Array.new(output_data)
        blob = Blob.new([excel_uint8], { "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" })
        url = URL.createObjectURL(blob)

        ### Create a download link and add style and turn it active
        download_link = document.getElementById("download-link")
        download_link.disabled = False
        download_link.style.cursor = "pointer"
        download_link.classList.remove("bg-gray-300", "text-gray-500", "cursor-not-allowed")
        download_link.classList.add("bg-blue-600", "hover:bg-blue-700", "text-white")

        def trigger_download(event):
            temp_link = document.createElement("a")
            temp_link.href = url
            temp_link.download = "processed_data.xlsx"
            temp_link.click()
            console.log("Download link created")

        download_link.addEventListener("click", create_proxy(trigger_download))
        download_link.click()
        console.log("Download button activated and triggered")
        


    #  Create proxy so Pyodide doesn’t destroy it
    reader.onload = create_proxy(combine)
    reader.readAsArrayBuffer(file)

# Also wrap the outer handler in a proxy
document.getElementById("excel-upload").addEventListener("change", create_proxy(allow_click))
document.getElementById("process-btn").addEventListener("click", create_proxy(handle_upload))
document.getElementById("remove-file-btn").addEventListener("click", create_proxy(remove_file))



